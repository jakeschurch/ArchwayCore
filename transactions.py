#!/usr/bin/python3
# -*- coding: future_fstrings -*-

from __future__ import absolute_import
import openpyxl
import os


class TransactionError(Exception):

    def __init__(self, ticker, qty, negAmt=None):
        self.ticker = ticker
        self.qty = qty
        self.negAmt = negAmt

    def __str__(self):
        if self.negAmt is not None:
            errStr = (f'Transacting {self.ticker} for {self.qty} shares '
                      f'creates a negative cash balance of {self.negAmt}. '
                      u'Please amend your transaction log.')
        else:
            errStr = (f'Cannot sell shares of {self.ticker} '
                      u'because none have been recorded as of yet in the '
                      u'transaction log. Please amend your transaction log.')
        return repr(errStr)


class Transaction(object):

    def __init__(self, **kwargs):
        self.ticker = kwargs[u'ticker']
        self.qty = float(kwargs[u'qty'])
        self.price = float(kwargs[u'price'].strip('$'))
        self.fees = float(kwargs[u'fees'])
        self.sector = kwargs[u'sector']
        self.action = kwargs[u'action']
        self.datetime = kwargs[u'datetime']

    def exportToPos(self):
        return {
            u'ticker': self.ticker,
            u'qty': self.qty,
            u'costBasis': self.price,
            u'sector': self.sector,
            u'buyDate': self.datetime,
            u'sellDate': None,
            u'costSold': None,
            u'dateSold': None
        }

    def exportToClosed(self):
        return {
            u'ticker': self.ticker,
            u'qty': self.qty,
            u'costSold': self.price,
            u'sector': self.sector,
            u'sellDate': self.datetime,
            u'costBasis': None,
            u'buyDate': None,
        }

    def __str__(self):
        print self.__dict__


def GetTx(txFile):
    txList = LoadTx(txFile, 0)
    cashTxList = GetCashTx(txList)
    allTx = txList + cashTxList

    return allTx


def GetCashTx(txList):
    cashTx = []
    for tx in txList:
        amt = tx.qty * tx.price
        if tx.Ticker != u"FDRXX" and amt != 0:
            txDetails = {
                u"ticker": u"FDRXX",
                u"qty": amt,
                u"price": 1.00,
                u"fees": None,
                u"action": u"Buy" if u"sell" in tx.action.lower() else u"Sell",
                u"datetime": tx.datetime
            }
            cashTx.append(Transaction(**txDetails))
    return cashTx


def LoadTx(df, startRow=0):
    txList = []
    for _, row in df.iterrows():
        rowPrice = row.loc[u'Sale Price/Share']
        if type(rowPrice) is unicode:
            rowPrice = float(rowPrice.replace(u'$', u''))

        rowShares = row.loc[u'Number of Shares']
        if type(rowShares) is unicode:
            rowShares = float(rowShares.replace(u',', u''))

        rowFees = row.loc[u'Fees']
        if type(rowFees) is unicode:
            rowFees = float(rowFees.replace(u'$', u''))

        kwargs = {
            u'action': row.loc[u'Action'],
            u'datetime': row.loc[u'Date Of Transaction'].to_pydatetime(),
            u'ticker': row.loc[u'Ticker'],
            u'sector': row.loc[u'Sector'],
            u'qty': rowShares,
            u'price': rowPrice,
            u'fees': rowFees
        }
        txList.append(Transaction(**kwargs))
    return txList


def OutputTx(fileName, txList):
    wb = openpyxl.load_workbook(
        os.path.abspath(u'data_files/template.xlsx')
    )
    ws = wb.copy_worksheet(wb[u'transactions'])

    i = 1
    for tx in txList:
        ws[u'a{0}'.format(i)] = tx.action
        ws[u'b{0}'.format(i)] = tx.ticker
        ws[u'c{0}'.format(i)] = tx.qty
        ws[u'd{0}'.format(i)] = tx.price
        ws[u'e{0}'.format(i)] = tx.fees
        ws[u'f{0}'.format(i)] = tx.sector
        ws[u'g{0}'.format(i)] = tx.datetime
    i += 1


if __name__ == u"__main__":
    LoadTx(None, 0)
