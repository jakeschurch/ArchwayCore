from __future__ import absolute_import
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Alignment, Font, Side, Border, PatternFill
from openpyxl.cell import Cell
from openpyxl.styles.Color import BLACK

# Copyright 2018 Jake Schurch
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


class styler(object):
    _metaheader = {
        u'alignment': Alignment(horizontal=True, wrap_text=True),
        u'font': Font(bold=True),
        u'fill': PatternFill(patternType=u'solid', fill_type=u'solid',
                             fgColor=u'D8A618'),
        u'border': Border(
            left=Side(border_style=u'normal', color=BLACK),
            right=Side(border_style=u'normal', color=BLACK),
            top=Side(border_style=u'normal', color=BLACK),
            bottom=Side(border_style=u'normal', color=BLACK),
            outline=Side(border_style=u'normal', color=BLACK),
            vertical=Side(border_style=u'normal',
                          color=BLACK),
            horizontal=Side(border_style=u'normal',
                            color=BLACK)
        )
    }
    _header = {
        u'alignment': Alignment(horizontal=True, wrap_text=True),
        u'font': Font(bold=True, color=u'FFD8A618'),
        u'fill': PatternFill(patternType=u'solid', fill_type=u'solid',
                             fgColor=BLACK),
        u'border': Border(
            left=Side(border_style=u'normal', color=BLACK),
            right=Side(border_style=u'normal', color=BLACK),
            top=Side(border_style=u'normal', color=BLACK),
            bottom=Side(border_style=u'normal', color=BLACK),
            outline=Side(border_style=u'normal', color=BLACK),
            vertical=Side(border_style=u'normal',
                          color=BLACK),
            horizontal=Side(border_style=u'normal',
                            color=BLACK)
        )
    }

    @staticmethod
    def metaheader(cell):
        cell.font = styler._metaheader[u'font']
        cell.fill = styler._metaheader[u'fill']
        cell.alignment = styler._metaheader[u'alignment']
        cell.border = styler._metaheader[u'border']

    @staticmethod
    def header(cell):
        cell.font = styler._header[u'font']
        cell.fill = styler._header[u'fill']
        cell.alignment = styler._header[u'alignment']
        cell.border = styler._header[u'border']
