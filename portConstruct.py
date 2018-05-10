#!/usr/bin/env python
# -*- coding: future_fstrings -*-

# Copyright 2018 Jake Schurch
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from __future__ import division
from __future__ import absolute_import
import openpyxl
import transactions as Tx
import datetime as dt
import numpy as np
import pandas as pd
import xlFuncs
import os


def readCSV(fileLoc, startRow=0):
    csv = pd.read_csv(
        fileLoc, header=startRow, skip_blank_lines=True, memory_map=True)
    csv.dropna(inplace=True, thresh=4)

    # strip whitespace from column headers
    csv.rename(columns=lambda x: x.strip(), inplace=True)

    # recode date column
    for header in csv.columns:
        if u'date' in header.lower():
            csv[header] = csv[header].apply(
                lambda x: pd.to_datetime(unicode(x), format=u'%m/%d/%Y'))

    # turn nan fee values into 0
    csv[u'Fees'].replace(np.nan, 0, inplace=True)

    # sort on date
    csv.sort_values(
        by=[u'Date Of Transaction', u'Action', u'Number of Shares'],
        ascending=[True, True, False],
        inplace=True)
    return csv


class PortfolioBuilder(object):
    def __init__(self):
        self.cash = 0
        self.activePos = {}
        self.closedPos = []

    def __str__(self):
        print u"Active Positions:"
        for pos in self.activePos:
            print f'\t{pos}'
            for tx in self.activePos[pos]:
                print f'\t\t{tx.__dict__}'

    def loadTransactions(
            self, transactions
            ):
       
        for tx in transactions:
            self.executeTx(tx)

    def executeTx(self, tx):
        if u'buy' in tx.action.lower():
            self._buy(tx)
        elif u'sell' in tx.action.lower():
            self._sell(tx)
        elif u'distribution' in tx.action.lower():
            self._distribute(tx)

    def _buy(self, tx):
        cashLeft = self.cash - (tx.price * tx.qty + tx.fees)
        if cashLeft < 0:
            print Tx.TransactionError(tx.ticker, tx.qty, cashLeft)

        self.cash = cashLeft  # Update cash position
        pos = Position(**tx.exportToPos())

        if tx.ticker not in self.activePos:
            self.activePos[tx.ticker] = []
        self.activePos[tx.ticker].append(pos)

        return pos

    def _sell(self, tx):
        aggQty = sum([pos.qty for pos in self.activePos[tx.ticker]])

        qtyLeftToSell = (aggQty - tx.qty)

        if qtyLeftToSell >= 0:
            self.cash += tx.price * tx.qty + tx.fees
            while tx.qty > 0:
                tx.qty = self.sellShares(self.activePos[tx.ticker], tx)
            return
        else:
            raise Tx.TransactionError(tx.ticker, tx.qty)

    def sellShares(self, posList, tx):
        pos = posList[0]

        if pos.qty > tx.qty:  # selling tx.qty shares
            qtyToSell = tx.qty
            tx.qty -= tx.qty
        else:
            qtyToSell = pos.qty
            tx.qty -= pos.qty
        leftOver = tx.qty

        soldPos = Position(**tx.exportToClosed())
        soldPos.qty = qtyToSell
        soldPos.buyDate = pos.buyDate
        soldPos.costBasis = pos.costBasis

        self.closedPos.append(soldPos)
        pos.qty -= qtyToSell

        if pos.qty == 0:
            del posList[0]
            # posList = posList[0:]
        return leftOver

    def getAggrQty(self, ticker):
        aggQty = 0
        for pos in self.activePos[ticker]:
            aggQty += pos.qty
        return aggQty

    def _distribute(self, tx):
        u""" NOTE: if stock distribution, cost basis is as of day before"""

        if u'cash' not in tx.ticker.lower() and u'fdrxx' not in tx.ticker.lower(
        ):
            self.activePos[tx.ticker].append()
        else:
            self._infuseCash(tx.qty)
        return

    def _infuseCash(self, amount):
        self.cash += amount

        return

    def getAggrPos(self, key):
        pos = None
        for index, iPos in self.activePos[key].items():

            if index == 0:
                pos = Position(iPos.exportData())
            else:
                pos.qty += iPos.qty
                pos.costBasis += iPos.costBasis

                if iPos.buyDate <= pos.buyDate:
                    pos.buyDate = iPos.buyDate

        pos.costBasis /= len(self.activePos[key])

        return pos

    def getBySector(self, sector=None):
        openPositions = []
        for k, _ in self.activePos.items():
            if k != []:
                if (self.activePos[k][0].sector == sector or
                        sector == u'All'):
                    openPositions.append(self.activePos[k])
        return openPositions


class Position(object):
    def __init__(self, **kwargs):
        self.ticker = kwargs[u'ticker']
        self.sector = kwargs[u'sector']

        self.qty = kwargs[u'qty']
        self.costBasis = kwargs[u'costBasis']
        self.costSold = kwargs[u'costSold']

        self.buyDate = kwargs[u'buyDate']
        self.sellDate = kwargs[u'sellDate']

    def exportData(self):
        return self.__dict__

    def __str__(self):
        return unicode(self.__dict__)


def WriteRow(writer, sheet, values):
    j = 1
    for v in values:
        sheet.cell(row=writer.i, column=j).value = v
        j += 1
    writer.i += 1

    return None


class currentWriter(object):
    u"""currentWriter writes data about current positions
        to an excel worksheet."""

    def __init__(
        self,
        port,
        sheet,
        sectorWanted,
        functioner,
        dateof,
        datefrom,
    ):
        self.port = port
        self.sheet = sheet
        self.functioner = functioner
        self.i = 0  # row index
        self.dateof = dateof
        self.datefrom = datefrom
        self.sector = sectorWanted

    def Write(self, sheet):
        # Write headers
        self._headers(sheet)

        # Write position data
        if self.sector == u'All':
            WriteRow(
                self,
                sheet,
                [
                    # First data block (A:J)
                    u'FDRXX',
                    u'Fidelity Government Cash Reserves Shs of Benef Interest',
                    u'=#N/A',
                    self.port.cash,
                    1.00,
                    self.port.cash,
                    1.00,
                    self.port.cash,
                    0,
                    0,
                    # Second data block (K:M)
                    u'',
                    1.00,
                    # TODO: get Beginning cash value
                    u'TODO',

                    # Third data block (N:Z)
                    u'', 0, 0, 0, u'=#N/A()',
                    # TODO: get weight of cash
                    u'',
                    0, 0, 0, 0,
                ]
            )
        data = self.port.getBySector(self.sector)
        for posList in data:
            self._data(sheet, posList)

        # Write total row
        self._total(sheet)

        return self.i

    def _headers(self, sheet):
        # Write meta-headers.
        sheet.merge_cells(u'A2:B2')
        sheet[u'A2'] = u'Holdings as of: {0}'.format(
            self.dateof.strftime(u'%m/%d/%Y'))

        sheet.merge_cells(u'G2:J2')
        sheet[u'G2'] = u'Current Value'

        sheet.merge_cells(u'O2:R2')
        sheet[u'O2'] = u'Dividends'

        sheet.merge_cells(u'L2:M2')
        sheet[u'L2'] = u'Beginning of Period'

        sheet.merge_cells(u'T2:W2')
        sheet[u'T2'] = u'Year to Date Performance'

        sheet.merge_cells(u'X2:Z2')
        sheet[u'X2'] = u'Beta'

        # Write headers.
        __headers = [
            # First Header Block (A:J)
            u'Ticker', u'Security Name', u'Date Acquired',
            u'Quantity', u'Avg Cost Basis',
            u'Cost Basis Total', u'Price', u'Market Value',
            u'HPR $', u'HPR %',

            # Second Header Block (K:M)
            u'', u'Price', u'Market Value',

            # Third Header Block (N:R)
            u'', u'Div Paid YTD', u'Total Dividends Collected YTD',
            u'Most Recent Dividend Payment', u'Div Date',

            # Fourth Header Block (S:Z)
            u'', u'YTD ($)', u'YTD (%)',
            u'Stock Weight in Sector', u'Contribution to YTD Sector Return',
            u'1 Year', u'3 Year', u'5 Year', u"ESG"
        ]
        self.i = 3
        return WriteRow(self, sheet, __headers)

    def _data(self, sheet, posList):
        u"""Write current holdings of a sector to excel sheet."""
        try:
            firstPos = posList[0]
        except IndexError:
            return

        aggrQty = self.port.getAggrQty(firstPos.ticker)
        # Write first block of values to excel sheet.
        row = [
            # First block of data --------------

            # Ticker
            firstPos.ticker,

            # Company Name
            u'=' + self.functioner.compName(firstPos.ticker),

            # Earliest Date Bought
            min([pos.buyDate for pos in posList]).strftime(u'%m/%d/%Y'),

            # Total Quantity
            sum([pos.qty for pos in posList]),

            # Avg Cost Basis
            sum([pos.costBasis * (pos.qty / aggrQty) for pos in posList]),

            # Total Cost Basis
            sum([pos.qty * pos.costBasis for pos in posList]),

            # Current Price
            u'=' + self.functioner.histPrice(firstPos.ticker),

            # Market Value
            f'=D{self.i}*G{self.i}',

            # HPR ($)
            f'==H{self.i}*F{self.i}',

            # HPR (%)
            f'==I{self.i}*F{self.i}',

            u'',
            # Second block of data --------------

            # Beginning of Period Price
            u'=' + self.functioner.histPrice(firstPos.ticker, self.datefrom),

            # Beginning of Period Market Value
            f'=L{self.i}*D{self.i}', u'',

            # Third block of data --------------

            # Divs Paid YTD
            u'=' + self.functioner.dividends(
                firstPos.ticker, startDate=self.datefrom),

            # Total Dividends Collected YTD
            u'=' + u'+'.join([self.functioner.dividends(pos)
                             for pos in posList]),

            # TODO:  Most Recent Dividend Payment
            u'',

            # TODO:  Most Recent Dividend Date
            u'', u'',

            # Fourth block of data --------------

            # YTD ($)
            f'=IFERROR(H{self.i}+P{self.i}-M{self.i}, \"\")',

            # YTD (%)
            f'=IFERROR((T{self.i}+P{self.i})/M{self.i}, \"\")',

            # TODO: stock weight in sector
            u'',

            # contribution to sector return ytd
            f'=IFERROR(U{self.i}*V{self.i}), \"\")',

            # Beta 1-Year
            u'=' + self.functioner.beta(
                firstPos.ticker, dateAsOf=self.dateof, betaYr=1),

            # Beta 3-Year
            u'=' + self.functioner.beta(
                firstPos.ticker, dateAsOf=self.dateof, betaYr=3),

            # Beta 5-Year
            u'=' + self.functioner.beta(
                firstPos.ticker, dateAsOf=self.dateof, betaYr=5),

            # ESG
            self.functioner.esg(firstPos.ticker),
        ]
        return WriteRow(self, sheet, row)

    def _total(self, sheet):
        sheet.merge_cells(f'A{self.i}:B{self.i}')
        sheet[f'A{self.i}'] = u"TOTALS"

        # Write column sums
        self.__writeTotals(sheet, u'fhimopqtuvw', 4, self.i)

        # Write other columns
        # 1yr beta
        sheet[f'X{self.i}'] = f'=sumproduct(v4:v{self.i-1}, x4:x{self.i-1})'

        # 3yr beta
        sheet[f'Y{self.i}'] = f'=sumproduct(v4:v{self.i-1}, y4:y{self.i-1})'

        # 5yr beta
        sheet[f'Z{self.i}'] = f'=sumproduct(v4:v{self.i-1}, z4:z{self.i-1})'

        return None

    def __writeTotals(self, sheet, cols, startRow, endRow):
        for col in cols:
            totalValue = f'=sum({col}{startRow}:{col}{endRow-1})'
            sheet[f'{col}{endRow}'] = totalValue
        return None


class realizedWriter(object):
    def __init__(
        self,
        closed,
        functioner,
        iLast,
        dateof,
        datefrom,
    ):
        self.iLast = iLast + 3
        self.closed = closed
        self.functioner = functioner
        self.i = iLast + 3  # row index
        self.dateof = dateof
        self.datefrom = datefrom

    def Write(self, sheet):

        # Write headers
        self._headers(sheet)

        # Write realized holding data
        self._data(sheet)

        # Write totals
        self._total(sheet)

        return self.i

    def _headers(self, sheet):
        # Write meta-headers
        asOf = f'Realized Holdings as of: {self.dateof.strftime("%m-%d-%Y")}'

        sheet.merge_cells(f'A{self.i}:B{self.i}')
        sheet[f'A{self.i}'] = asOf

        sheet.merge_cells(f'F{self.i}:G{self.i}')
        sheet[f'F{self.i}'] = u'Beginning of Year'

        sheet.merge_cells(f'H{self.i}:I{self.i}')
        sheet[f'H{self.i}'] = u'At Sale'
        self.i += 1

        # Write Headers
        __headers = [
            u'Ticker',
            u'Security Name',
            u'Date Acquired',
            u'Date Sold',
            u'Quantity',
            u'Price',
            u'Market Value',
            u'Price',
            u'Market Value',
            u'Dividends Collected',
            u'YTD ($)',
            u'YTD (%)',
        ]
        return WriteRow(self, sheet, __headers)

    def _data(self, sheet):
        for pos in self.closed:
            WriteRow(self, sheet, self.__outputPos(pos))

        return None

    def __outputPos(self, pos):
        divs = self.functioner.dividends(pos.ticker, pos.buyDate, pos.sellDate)
        return [
            pos.ticker,
            u'=' + self.functioner.compName(pos.ticker),
            pos.buyDate.strftime(
                u'%m/%d/%Y'), pos.sellDate.strftime(u'%m/%d/%Y'),
            pos.qty, pos.costBasis, f'=d{self.i}*e{self.i}',
            pos.costSold, f'=f{self.i}*e{self.i}',
            f'={divs}*e{self.i}', f'=(i{self.i}+j{self.i})-g{self.i}',
            f'=K{self.i}/G{self.i}',
        ]

    def _total(self, sheet):
        sheet[f'A{self.i}'] = u'TOTALS'
        self.__writeTotals(sheet, u'gijk', self.iLast, self.i)

    def __writeTotals(self, sheet, cols, startRow, endRow):
        for col in cols:
            totalValue = f'=sum({col}{startRow}:{col}{endRow-1})'
            sheet[f'{col}{endRow}'] = totalValue

        return None


class alphaWriter(object):
    def __init__(
        self,
        currentEndRow,
        realizedEndRow,
    ):
        self.iCurrent = currentEndRow
        self.iRealized = realizedEndRow
        self.i = realizedEndRow + 4

    def Write(self, sheet):
        self._headers(sheet)
        self._data(sheet)
        self._total(sheet)

        return None

    def _headers(self, sheet):
        # Write meta-headers
        sheet.merge_cells(f'A{self.i}:B{self.i}')
        sheet[f'A{self.i}'] = u'Year to Date Alpha'

        sheet.merge_cells(f'C{self.i}:G{self.i}')
        sheet[f'C{self.i}'] = u'Beginning of Period/Purchase'

        sheet.merge_cells(f'F{self.i}:G{self.i}')
        sheet[f'F{self.i}'] = u'Current Value/Sale'
        self.i += 1

        headers = [
            u'Ticker',
            u'Security Name',
            u'Realized',
            u'Market Value',
            u'Weight',
            u'Market Value',
            u'Weight',
            u'YTD ($) w/ Dividends',
            u'YTD (%) w/ Dividends',
            u'Weighted Contribution to Return',
            u'',
            u'YTD', u'Benchmark', u'Alpha',
        ]
        return WriteRow(self, sheet, headers)

    def _data(self, sheet):
        # Write Current Data
        currentStart = 4
        for i in xrange(currentStart, self.iCurrent):
            WriteRow(
                self,
                sheet,
                [
                    f'=A{i}',
                    f'=B{i}',
                    u'',
                    f'=M{i}',
                    # TODO: beginning weight
                    u'',
                    f'=H{i}',
                    # TODO:  current market weight
                    u'',
                    f'=T{i}',
                    f'=H{self.i}/D{self.i}',
                    f'=G{self.i}/I{self.i}',
                ]
            )

        # Write realized data
        realizedStart = self.iCurrent + 6
        for i in xrange(realizedStart, self.iRealized):
            WriteRow(
                self,
                sheet,
                [
                    f'=A{i}',
                    f'=B{i}',
                    u'',
                    f'=M{i}',
                    # TODO: beginning weight
                    u'',
                    f'=H{i}',
                    # TODO:  current market weight
                    u'',
                    f'=T{i}',
                    f'=H{self.i}/D{self.i}',
                    f'=G{self.i}/I{self.i}',
                ]
            )

        return None

    def _total(self, sheet):
        alphaStart = self.iRealized + 6

        # Write total row.
        sheet[f'A{self.i}'] = u'TOTALS'
        WriteRow(
            self,
            sheet,
            [
                u'',
                u'',
                u'',
                f'=sum(D{alphaStart}:D{self.i-1})',
                f'=sum(E{alphaStart}:E{self.i-1})',
                f'=sum(F{alphaStart}:F{self.i-1})',
                f'=sum(G{alphaStart}:G{self.i-1})',
                f'=sum(H{alphaStart}:H{self.i-1})',
                f'=H{self.i}/D{self.i}',
            ]
        )
        # Write relative return data
        sheet[f'L{alphaStart}'] = f'=J{self.i-1}'
        # TODO: relative return against benchmark
        sheet[f'M{alphaStart}'] = u'TODO:'
        sheet[f'N{alphaStart}'] = f'=L{alphaStart}-M{alphaStart}'

        return None


class ExcelWriter(object):

    def __init__(
            self,
            port,
            sectorsWanted=[u'All'],
            funcsWanted=u'factset',
            endDate=dt.datetime.today(),
            startDate=dt.datetime.today() - dt.timedelta(days=365)
    ):
        self.sectors = sectorsWanted

        self.rowIndex = 0
        self.Functioner = xlFuncs.xlFunctionSelector(startDate, endDate,
                                                     funcsWanted)
        self.port = port
        self.endDate = endDate
        self.startDate = startDate
        self.realizedIndex = 4

    def Write(self, outputPath):
        templatePath = os.path.abspath(u'data_files/template.xlsx')
        wb = openpyxl.load_workbook(templatePath)

        for sector in self.sectors:
            sectorSheet = wb.copy_worksheet(wb[u'Sheet1'])
            sectorSheet.title = sector if sector != u'All' else u'Archway Fund'

            # Write Current Holding Data
            current = currentWriter(
                self.port, sectorSheet, sector, self.Functioner,
                self.startDate, self.endDate
            )
            iCurrentEnd = current.Write(sectorSheet)

            # Write Realized Holding Data
            realized = realizedWriter(
                self.port.closedPos, self.Functioner,
                iCurrentEnd, self.startDate, self.endDate,
            )
            iRealizedEnd = realized.Write(sectorSheet)

            # Write YTD Alpha Data
            alpha = alphaWriter(
                iCurrentEnd, iRealizedEnd,
            )
            alpha.Write(sectorSheet)

        wb.remove_sheet(
            wb.get_sheet_by_name('Sheet1')
        )

        wb.save(u'portOutput.xlsx')
        # TODO: option to specify output path
