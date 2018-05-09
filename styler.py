from __future__ import absolute_import
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, Alignment, Font, Side, Border, PatternFill
from openpyxl.cell import Cell
from openpyxl.styles.Color import BLACK


class styler(object):
    _metaheader = {
        u'alignment': Alignment(horizontal=True, wrap_text=True),
        u'font': Font(bold=True),
        u'fill': PatternFill(patternType=u'solid', fill_type=u'solid',
                            fgColor=u'D8A618'),
        u'border': Border(
            left=Side(border_style=u'normal', color=BLACK),
            right=Side(border_style=u'normal', color=BLACK),
            top=Side(border_style=u'normal', color=BLACK),
            bottom=Side(border_style=u'normal', color=BLACK),
            outline=Side(border_style=u'normal', color=BLACK),
            vertical=Side(border_style=u'normal',
                          color=BLACK),
            horizontal=Side(border_style=u'normal',
                            color=BLACK)
        )
    }
    _header = {
        u'alignment': Alignment(horizontal=True, wrap_text=True),
        u'font': Font(bold=True, color=u'FFD8A618'),
        u'fill': PatternFill(patternType=u'solid', fill_type=u'solid',
                            fgColor=BLACK),
        u'border': Border(
            left=Side(border_style=u'normal', color=BLACK),
            right=Side(border_style=u'normal', color=BLACK),
            top=Side(border_style=u'normal', color=BLACK),
            bottom=Side(border_style=u'normal', color=BLACK),
            outline=Side(border_style=u'normal', color=BLACK),
            vertical=Side(border_style=u'normal',
                          color=BLACK),
            horizontal=Side(border_style=u'normal',
                            color=BLACK)
        )
    }

    @staticmethod
    def metaheader(cell):
        cell.font = styler._metaheader[u'font']
        cell.fill = styler._metaheader[u'fill']
        cell.alignment = styler._metaheader[u'alignment']
        cell.border = styler._metaheader[u'border']

    @staticmethod
    def header(cell):
        cell.font = styler._header[u'font']
        cell.fill = styler._header[u'fill']
        cell.alignment = styler._header[u'alignment']
        cell.border = styler._header[u'border']
